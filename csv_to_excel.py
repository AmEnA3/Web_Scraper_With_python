

import pandas as pd
from pathlib import Path


def convert_csv_to_excel():
    """
    Convert CSV file to Excel format with proper formatting.
    
    Reads the CSV file created by the scraper and converts it to an Excel
    file with improved formatting for better readability.
    """
    
    # Define file paths
    csv_file = "univ_eloued_activities.csv"
    excel_file = "university_activities.xlsx"
    
    print("=" * 70)
    print("CSV to Excel Converter")
    print("=" * 70)
    
    try:
        # Check if CSV file exists
        if not Path(csv_file).exists():
            print(f"Error: {csv_file} not found in current directory.")
            print("Please run the scraper first: python univ_eloued_scraper.py")
            return
        
        print(f"\nReading CSV file: {csv_file}")
        
        # Read CSV file with semicolon separator
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8')
        
        # Display basic info about the data
        print(f"  • Records loaded: {len(df)}")
        print(f"  • Columns: {', '.join(df.columns.tolist())}")
        print(f"  • Data shape: {df.shape}")
        
        print(f"\nConverting to Excel format...")
        
        # Write DataFrame directly to Excel file using openpyxl engine
        # This is the main conversion using df.to_excel()
        df.to_excel(excel_file, sheet_name='Activities', index=False, engine='openpyxl')
        
        print(f"  ✓ Excel file written: {excel_file}")
        
        # Now load the workbook to apply additional formatting
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            # Format the header row with blue background and white text
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=12
            )
            
            # Apply formatting to header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"  ✓ Header formatting applied")
            
            # Adjust column widths for better readability
            # New CSV layout: Title, Date, Description, Link
            column_widths = {
                'A': 60,   # Title
                'B': 18,   # Date
                'C': 100,  # Description
                'D': 60    # Link
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            print(f"  ✓ Column widths adjusted")
            
            # Enable text wrapping for all data cells
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
            
            # Set row heights for better appearance
            worksheet.row_dimensions[1].height = 25  # Header row
            for row_num in range(2, len(df) + 2):
                worksheet.row_dimensions[row_num].height = 60  # Data rows
            
            print(f"  ✓ Text wrapping enabled")
            
            # Save the formatted workbook
            workbook.save(excel_file)
            print(f"  ✓ Formatting saved")
            
        except ImportError:
            print(f"  ⚠ Warning: openpyxl formatting not applied")
            print(f"    (Excel file created, but without enhanced formatting)")
        
        # Display file info
        file_size = Path(excel_file).stat().st_size
        print(f"\nOutput file: {excel_file}")
        print(f"File size: {file_size:,} bytes")
        print(f"Records: {len(df)}")
        
        print("\n" + "=" * 70)
        print(" Conversion completed successfully!")
        print("=" * 70)
        print(f"You can now open '{excel_file}' in Excel or LibreOffice Calc")
        
    except ImportError as e:
        print(f"\n Error: Missing required library")
        print(f"   {str(e)}")
        print("\nInstall the required libraries:")
        print("   pip install pandas openpyxl")
        
    except Exception as e:
        print(f"\n Error: {str(e)}")
        print("Please check that the CSV file exists and is properly formatted.")


if __name__ == '__main__':
    convert_csv_to_excel()
